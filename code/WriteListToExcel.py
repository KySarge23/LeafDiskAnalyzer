#!/usr/bin/python3

'''
    this is a method to write a list of data to an excel document
    need the spreadsheet to be created before the method is called
    Excel sheet cannot be open while running (gives permission denied error)
'''

from openpyxl import load_workbook


#function for putting lists into columns
def putValuesIntoExcelColumn(values, workbookName = 'testingExcel.xlsx',sheetName = 'Sheet1', columnNum = 1, rowNum = 1):
    
    wb = load_workbook(workbookName)
    ws = wb[sheetName]
    for ratio in values:
        #ws['A1'] = i
        d = ws.cell(row = rowNum, column= columnNum, value= ratio)
        rowNum += 1

    
    wb.save('testingExcel.xlsx')

#function for putting lists into rows
def putValuesIntoExcelRow(values, workbookName = 'testingExcel.xlsx',sheetName = 'Sheet1', columnNum = 1, rowNum = 1):
    
    wb = load_workbook(workbookName)
    ws = wb[sheetName]
    for ratio in values:
        #ws['A1'] = i
        d = ws.cell(row = rowNum, column= columnNum, value= ratio)
        columnNum += 1

    
    wb.save('testingExcel.xlsx')


def main():
    #Example list of ratios
    ratios = [.93, .94, .45, .89, .33]

    #This funtion can take 5 vaules "list of Data", "Starting Column", "Starting Row"
    #first value is required, second 2 "workbookName" and "sheetName" should be required,
    #but for now have them set to a testing spreadsheet
    #if last two values are not giving the info start at row 1, column 1 by default

    #testing columns function
    putValuesIntoExcelColumn(ratios)

    #testing rows function
    putValuesIntoExcelRow(ratios, 'testingExcel.xlsx', 'Sheet1', 2, 1)

main()
