     #!/usr/bin/env python3

#Imports that dont need to be installed via pip
import math
import os
import glob
import time
import tkinter as tk
import threading as thr
import numpy as np

#--------------------------------------------#

#Imports that need to be installed via pip:
import cv2 as cv #opencv-python
import openpyxl, openpyxl.styles #openpyxl

#--------------------------------------------#

#GUI Imports
import GUI
import calendarPicker as cp

#--------------------------------------------#

#Specific functionality imports
from pathlib import Path as pth
from tkinter.filedialog import askopenfilename, asksaveasfilename
from tkinter import messagebox

#--------------------------------------------#


# Author(s): Kyle Sargent, Connor Jansen, Colton Eddy, Alex Wilson, Emily Box
# Version: 1
# Copyright: Copyright (c) 2019, Missouri State University
# All rights reserved.
# Redistribution and use in source and binary forms, with or without modification, are permitted provided that the
# following conditions are met:
# * Redistributions of source code must retain the above copyright notice, this list of conditions and the following
# disclaimer.
# * Redistributions in binary form must reproduce the above copyright notice, this list of conditions and the
# following disclaimer in the documentation and/or other materials provided with the distribution.
# * Neither the name of the organization nor the names of its contributors may be used to endorse or promote
# products derived from this software without specific prior written permission.
# THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS" AND ANY
# EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED WARRANTIES OF
# MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE DISCLAIMED. IN NO EVENT
# SHALL MISSOURI STATE UNIVERSITY BE LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL,
# EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF
# SUBSTITUTE GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION)
# HOWEVER CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY,
# OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS
# SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE."
# -------------------------------------------------------------------------------------------------------------------

thrLock = thr.Lock()

def writeToExcel(value, workbookName, sheet, date, picNum,): #phenoNum):
    """

    This function will handle any writing to the spreadsheet document that is required.
    It uses openpyxl to load in a workbook and write the data obtained from the analysis to spreadsheets in selected workbooks.

    Input(s): value (int), workbookName (string), sheet (string), date (string), picNum (int)
    Output(s): Data written to sheet in workbook
    Local Variable(s):  wb (excel workbook that is opened), ws (worksheet from opened workbook), startCol/startRow (int), 
    
    """

    wb = openpyxl.load_workbook(workbookName)
    ws = wb[sheet]
    startCol, startRow = 1, 1
    dateCol, dateRow = 0, 0
    foundDate = False
    pic = "Picture: " + str(picNum)

    for cell in ws[1]:
        if cell.value == date:
            dateCol = cell.column
            dateRow = cell.row
            foundDate = True
            break
    
    if foundDate == True:
        ratioRow = picNum + 1
        ratioCol = dateCol
        ratioCell = ws.cell(ratioRow , ratioCol, value = pic + ", Ratio: "+ str(value))
        wb.save(workbookName)
        return

    elif foundDate == False:
        if ws.cell(startRow,startCol).value == None:
            dateCell = ws.cell(startRow, startCol, value = date)
            dateCell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
            dateCol = startCol
            dateRow = startRow
            ws.merge_cells(dateCell.coordinate + ':' + ws.cell(dateRow, dateCol + 3).coordinate)
            wb.save(workbookName)

        else:
            for i in range(startCol, 200, 6):
                if ws.cell(row = startRow, column = i).value != None: continue
                else: 
                    dateCell = ws.cell(startRow, i, value = date)
                    dateCell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
                    dateCol = i
                    dateRow = startRow
                    ws.merge_cells(dateCell.coordinate + ':' + ws.cell(dateRow, dateCol + 3).coordinate)
                    wb.save(workbookName)
                    break

        ratioRow = picNum + 1
        ratioCol = dateCol
        ratioCell = ws.cell(ratioRow , ratioCol, value = pic + ", Ratio: "+ str(value))

    wb.save(workbookName)
    return

def backgroundRemove(img):
    height, width = img.shape[:2]

    #Create a mask holder
    mask = np.zeros(img.shape[:2],np.uint8)

    #Grab Cut the object
    bgdModel = np.zeros((1,65),np.float64)
    fgdModel = np.zeros((1,65),np.float64)

    #create Rect The object must lie witin
    rect = (2,0,width-5,height)
    cv.grabCut(img,mask,rect,bgdModel,fgdModel,4,cv.GC_INIT_WITH_RECT)
    mask = np.where((mask==2)|(mask==0),0,1).astype('uint8')
    img1 = img*mask[:,:,np.newaxis]

    #Get the background
    background = img - img1

    #Change all pixels in the background to what ever color you want
    #I went with Red because of high contast with the green leaf disk
    background[np.where((background > [0,0,0]).all(axis = 2))] = [0,0,0]

    #Add the background and the image
    final = background + img1
    return final

def calculateMildew(path):
    """
    
    Area finding function, Will utilize HoughCircle method to detect circles
    within the photo or hard coded method to detect circles. After finding the circle, the image is cropped down and then hsv detection is ran on 
    it. After the hsv is completed, Edge Detection is ran on the newly hsv filtered photo. After Edge detection, contouring is performed for  

    Input(s): path (String)
    Output(s): mildewRatio (int)
    Local Variable(s): img/cimg/hsv/mask/res/edges (numpy ndarray), height/width/area/rad/channel/mildewRatio/cont/contArea (int), hsvValues (array of ints), 
                       lower_green/upper_green/contours/hierarchy (nparray

    """

    img = cv.imread(path,1) 
    h,w = img.shape[:2]

    if h < 280 and w < 423:
        print("Image Resolution too small for analyzing! Image passed is too small for analyzing, please resize or retry with a different image.")
        return -1
    if h > 280 and w > 423:
        img = cv.resize(img,(423,280)) #resize image, for easier reading and faster execution.
   
    img = cv.medianBlur(img,5) #add blur to reduce noise on photo.
    cimg = cv.cvtColor(img, cv.COLOR_BGR2GRAY)
    
    img = backgroundRemove(img)

    center = (int(w / 2), int(h / 2))
    rad = 205
    cv.circle(cimg, center, rad, (0,0,255), 2)

    area = math.pi * rad ** 2 

    img = img[0:h, 0:w]

    hsv = cv.cvtColor(img, cv.COLOR_BGR2HSV)

    height, width, channel= hsv.shape

    #adding hsv values to list
    hsvValues = []
    for x in range(0, width):
        for y in range(0, height):
            pixel= hsv[y, x]
            #print(pixel)
            hsvValues.append(pixel)

    #darker colors (to mask out)
    lower_green = np.array([0,0, 165])
    #lighter colors
    upper_green = np.array([255, 255, 255])

    #masking pixels to remove irrelevant data
    mask = cv.inRange(hsv, lower_green, upper_green)
    res = cv.bitwise_and(img, img, mask=mask)

    hsvMask = cv.inRange(hsv, lower_green, upper_green)
    res = cv.bitwise_and(img, img, mask=hsvMask)

    edges = cv.Canny(res,350,700)

    ret, edges = cv.threshold(edges,130,255,cv.ADAPTIVE_THRESH_MEAN_C)
    mask = np.zeros(edges.shape,np.uint8)

    contours, hierarchy = cv.findContours(edges, cv.RETR_LIST, cv.CHAIN_APPROX_SIMPLE)

    mildewRatio = 0
    for cont in contours:
        contArea = cv.contourArea(cont)
        mildewRatio += contArea

    return mildewRatio / area * 100


def threadHandler(date, trayNum, picNum, spreadsheet):
    """
    
    This Function accepts the three user inputs from the main function/GUI as arguments.
    It then builds a valid filepath based on the arguments. Once a valid path is created,
    this function sends that path to the findCircleArea and cannyEdgeDetection functions. 
    This function validates the file extension of the file selected to make sure that the selection is a valid image format.
    Supported file extensions: .png, .jpeg, .jpg, .tiff, .tif

    Input(s): date (String) , tray (int), picNum (int)
    Output(s): None
    Local Variable(s): path (String), dirName (String) , fName (String), circArea (int), mildewRatio (int), mildewRatio (int)

    """

    print(thr.current_thread())

    date = glob.glob("../photos/"+ date + "*", recursive=True)[0]
    tray = '/tray '+ str(trayNum)
    dirName = date + tray + "/"
    try:
        photo = glob.glob(dirName + str(picNum) +"-160x271_"+ "*", recursive=True)[0]
    except:
        print("Path cannot be found! A path cannot be found for photo: " + str(picNum) + " in tray: " + str(trayNum) + " from the date: " + date[10:] +".")
        return

    path = os.path.abspath(photo)
    
    if os.path.exists(path):
        startTime = time.perf_counter()
        mildewRatio = calculateMildew(path)
        endTime = time.perf_counter()
        totalTime = endTime - startTime
        
        if mildewRatio < -1:
            return

        if totalTime > 10: return print("Timeout in: " + thr.current_thread().getName() + " with runtime of: " +str(totalTime) +"s.")
      
        thrLock.acquire()
        writeToExcel(mildewRatio, spreadsheet, tray[1:], date[10:], picNum)
        print(thr.current_thread().getName() + " returning.")
        print("Mildew to leaf ratio is: " + str(mildewRatio) + "%")
        thrLock.release()
        return 

    else:
        print("Invalid path detected, No file or directory resides in: \n" + path)
        return

def findOccurrences(s, ch):
    """
    
    Function to find occurances of a character in a string. only useful for when ',' is used in an entry field.

    Input(s): s (String), ch (Character)
    Output(s): a list of indicies in which ch was found in s.
    Local Variable(s): None

    """    
    return [i for i, letter in enumerate(s) if letter == ch]

def getNumbers(input):
    """

    Function to extract the numbers from the entry fields after they've been validated. 
    This will grab the numbers found surrounding a '-' or multiple ',' in an entry field. 
    Then will place either the range of numbers or sequence of numbers in a list
    and return that to the Analyzer.

    Input(s): input (String)
    Output(s): nums[] (list of ints)
    Local Variables: idx (int), n1/n2 (int), nums[] (list of ints), comms [] (list of indicies where ',' is found)

    """

    nums, comms = [], []

    if '-' in input and ',' in input:
        comms = findOccurrences(input, ',')
        splitInput = input.split(',',len(comms))
        for splitStr in splitInput:
            if '-' in splitStr:
                splitSubStr = splitStr.split("-", 2)
                n1 = splitSubStr[0]
                n2 = splitSubStr[1]

                if n1  == '':
                    return nums #return nums here because invalid format found. do error handling in the sendToAnalyzer function.
                elif n2 == '':
                    return nums #same as for n1.

                n1 = int(n1)
                n2 = int(n2)
                if n1 > n2: 
                    return messagebox.showwarning("Entry Warning!", "Left Hand Side of '-' is greater than Right Hand Side. Please fix the order and retry.")
                else:
                    for i in range (n1, n2+1):
                        nums.append(i)
            else:
                n = int(splitStr)
                if n in nums:
                    continue
                else:
                    nums.append(n)
        return nums
        
    
    if '-' in input:
        splitInput = input.split("-", 2)
        n1 = splitInput[0]
        n2 = splitInput[1]
        if n1  == '':
            return nums #return nums here because invalid format found. do error handling in the sendToAnalyzer function.
        elif n2 == '':
            return nums #same as for n1.

        n1 = int(n1)
        n2 = int(n2)
        if n1 > n2: 
            return messagebox.showwarning("Entry Warning!", "Left Hand Side of '-' is greater than Right Hand Side. Please fix the order and retry.")
        else:
            for i in range (n1, n2+1):
                nums.append(i)
            return nums
     
    if ',' in input and len(input) >= 3:
        comms = findOccurrences(input, ',')
        splitInput = input.split(',', len(comms))
        for splitNum in splitInput:
            n = int(splitNum)
            if n in nums:
               continue
            else:
                nums.append(n)
        return nums

    if '-' not in input and ',' not in input:
        n1 = int(input)
        nums.append(n1)
        return nums


def date():
    global datePicker
    datePicker = cp.DatePicker(format_str="%01d-%02d-%02d")
    

def returnDate():
    return datePicker.date

def main():
    root = tk.Tk()

    root.geometry('350x300') #set size of window
    root.title("LDA GUI v1.0")

    root.resizable(False, False) #make window not able to be resized

    def exitWindow(e):
        if messagebox.askyesnocancel("Exit","Are you sure you want to close this application?"):
            root.destroy()

    root.bind("<Escape>", exitWindow)
    gui = GUI.analyzerGUI(root) #create new instance of the analyzerGUI with root as master.

    trays, pics, threads = [], [], []
    workbook = ""

    gui.calendarBtn.config(command=date)


    def newOrExisting(trays):
        """
        
        Function for determining whether a new spreadsheet or an existing spreadsheet will be used to hold data from the analyzing process.
        We get the option selected from the radio buttons on the GUI and return True if one was selected, otherwise we show a 
        messagebox warning letting the user know neither button was picked, and return false

        Input(s): trays (list of trays gained from user input)
        Output(s): True or False Boolean
        Local Variable(s): value (String), new (Boolean)

        """
        value = gui.option.get()
        if value == "True":
                print("Creating New Spreadsheet")
                wb = openpyxl.Workbook()
                rm = wb['Sheet']
                wb.remove(rm)

                for tray in trays:
                    ws = wb.create_sheet("tray " + str(tray))
                
                file = asksaveasfilename(initialdir = ".",title = "Save As",filetypes = (("xlsx","*.xlsx"),("All Files","*.*")), defaultextension = '.xlsx',) #creates new workbook (currently creates a single placeholder book in the current directory               
                wb.save(file)
                return file

        elif value == "False":
                print("Going to Existing")
                file = askopenfilename(initialdir = ".", title = "Open File", filetypes = (("xlsx","*.xlsx"),("All Files","*.*"))) #lets user browse for what spreadsheet they want to open
                wb = openpyxl.load_workbook(file)
                wsheets = wb.sheetnames
                for tray in trays:
                    if "tray " + str(tray) not in wsheets:
                        ws = wb.create_sheet("tray " + str(tray))

                wb.save(file)
                return file

    def validateTP(x,y):
        """

        Function to validate the Tray/Picture entries. We initialize countx and county for counting correct or 
        valid characters found from the inputs.
        Then we loop over each string and check if the character is valid (e.g. digit, '-' or ','). 
        If so we increment count, if not we show a warning with the invalid character 
        and clear the entry field in which the invalid character was found. 
        After character validation, we check if countx and county 
        are equal to the length of the strings passed in, meaning that each character was valid.

        Input(s): x (String), y (String)
        Output(s): boolean value based on whether or not both strings are valid.
        Local Varaible(s): countx (int), county (int)

        """

        if x == "" and y == "":
            messagebox.showwarning("No Entry Warning!", "No entry found in the Tray and Picture Entry Fields. Please enter data in the following format: '1-3' or '1,2,3'.")
            return False

        elif x == "":
            messagebox.showwarning("No Entry Warning!", "No entry found in the Tray Entry Field. Please enter data in the following format: '1-3' or '1,2,3'.")
            return False
        elif y == "":
            messagebox.showwarning("No Entry Warning!", "No entry found in the Picture Entry Field. Please enter data in the following format: '1-3' or '1,2,3'.")
            return False

        countx, county = 0, 0
        for i in x:
            if i.isdigit() or i == '-' or i == ',':
                countx += 1
            else:
                messagebox.showwarning("Entry warning!", "Incorrect character of: '" + i + "' in Tray Entry Field. Please enter in the following format: '1-3' or '1,2,3'.")
                gui.trayEntry.delete(0,tk.END)                

                return False     
        for o in y:
            if o.isdigit() or o == '-' or o == ',':
                county += 1
            else: 
                messagebox.showwarning("Entry warning!", "Incorrect character of: '" + o + "' in Picture Entry Field. Please enter in the following format: '1-3' or '1,2,3'.")
                gui.picEntry.delete(0,tk.END)                
                return False

        if countx == len(x) and county == len(y):
            return True

    def validateDate(date): 
        """
        Function to validate the date input by the user. We check if date is less than 6 because it allows for folders to be structured 
        as d-mm-yy as well.
        
        Input(s): date (String)
        Output(s): path found from global search with date
        Local Variable(s):

        """
        dateLen = len(date)
        count = 0

        for i in date:
            if i.isdigit() or i == '-':
                count+=1
            else: break

        if count == dateLen:
            try: return glob.glob("../photos/"+ date + "*", recursive=True)[0]
            except: messagebox.showerror("No Associated Folder Found!", "No Folder was found for the date of: " + date + " please try again.")
        else:
            messagebox.showwarning("Incorrect Date Found", "An incorrect date of: " + date + " has been found. Please retry with the following format: mm-dd-yy.")
            return

    def disableAll(gui):
        gui.trayEntry.config(state = 'disabled')
        gui.picEntry.config(state = 'disabled')
        gui.calendarBtn.config(state = 'disabled')
        gui.r1.config(state = 'disabled')
        gui.r2.config(state = 'disabled')
        return

    def enableAll(gui):
        gui.trayEntry.config(state = 'normal')
        gui.picEntry.config(state = 'normal')
        gui.calendarBtn.config(state = 'normal')
        gui.r1.config(state = 'normal')
        gui.r2.config(state = 'normal')
        return


    def sendToAnalyzer():
        """
        Function to send data from GUI fields to diskAnalyzer. We grab the GUI fields' values and strip any 
        whitespace from the front/back so that it doesnt mess up with our validation methods. 
        Then we validate the fields and upon them returning true, 
        we disable all buttons and then run the analyzer methods with the validated data gained.
        
        Input(s): None
        Output(s) None
        Local Varaible(s): trayStr/picStr/dateStr (str), numTrays/numPics (int), t (thread obj), threads (list of threads), t/elapsedTime (float)

        """

        disableAll(gui)

        trayStr = gui.trayEntry.get().rstrip().lstrip().replace(" ", "")
        picStr = gui.picEntry.get().rstrip().lstrip().replace(" ", "")
        date = returnDate().rstrip().lstrip()

        print("Selected date: " + date)
        
        if validateTP(trayStr, picStr) and validateDate(date):
            trays = getNumbers(trayStr)
            pics = getNumbers(picStr)
        
            numTrays = len(trays)
            numPics = len(pics)

            if numPics  == 0 and numTrays == 0:
                enableAll(gui)
                return messagebox.showerror("Invalid Format Detected!", "An Invalid format was found in both the picture and tray numbers entry fields. Please retry with the following format: '1-3' or '1,2,3'.")
            elif numTrays == 0:
                enableAll(gui)
                return messagebox.showerror("Invalid Format Detected!", "An Invalid format was found in the tray numbers entry field. Please retry with the following format: '1-3' or '1,2,3'.")
            
            elif numPics == 0:
                enableAll(gui)
                return messagebox.showerror("Invalid Format Detected!", "An Invalid format was found in the tray numbers entry field. Please retry with the following format: '1-3' or '1,2,3'.")

            try:
                workbook = newOrExisting(trays)
            except:
                enableAll(gui)
                return messagebox.showerror("Selection Cancelled!", "Selection of existing sheet cancelled. Please try again.")    

            if workbook == ".xlsx": #if saving the workbook filename is cancelled, file will become ".xlsx" and still create a spreadsheet and run the analyzing. This will prevent that.
                os.remove(workbook) 
                enableAll(gui)
                return messagebox.showwarning("No Save/Existing Spreadsheet Name Warning!", "No name has been selected for the spreadsheet you wish to use. Please retry.")


            if numPics * numTrays > 8:
                enableAll(gui)
                return messagebox.showwarning("Input Warning!", "Current inputs from Tray/Pictures entry fields will spawn too many threads. Use the following as a guide for entering data into tray/pictures entry fields: trays * pictures <= 8.")
        
            messagebox.showinfo("Successful Verification!", "Data has been successfully verified. The analyzing process will now begin.")

            for i in range(len(trays)):
                for j in range(len(pics)):
                    t = thr.Thread(target = threadHandler, args = [date, trays[i], pics[j], workbook]) 
                    threads.append(t)
                    t.start()

            for  t in threads:
                t.join()
        
           
            
        enableAll(gui)
    
        print("Analyzing Complete.")
        return
                
    
    uploadBtn = tk.Button(root, text= "Analyze", command=sendToAnalyzer, height = 1, width = 10)
    uploadBtn.grid(row= 5, column = 0)    

    root.mainloop()


main()
