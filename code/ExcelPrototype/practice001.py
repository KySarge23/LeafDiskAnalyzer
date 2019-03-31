from openpyxl import Workbook, load_workbook
from tkinter.filedialog import askopenfilename
#import os


date = "09-16-18"
tray = 2
pic = 4
result = .322786221
newWS = True

cellRow = 1
cellCol = 2

cCol = 1
cRow = 1+pic

sName = ""
if newWS:
    sName = input("Input a name for the new spreadsheet: ") + '.xlsx'
    wb = Workbook()
    wb.active.title = "Tray 1"
    for i in range(1,8):
        wb.create_sheet("Tray " + str(i+1))
else:
    
    sName = askopenfilename()
    print(sName)
    #fName = os.path.basename(file)
    wb = load_workbook(str(sName))


for sheet in wb:
    
    c = sheet.cell(row = cellRow, column = cellCol)
    while c.value != None:
        cellCol += 1
        c = sheet.cell(row = cellRow, column = cellCol)
    c.value = date

ws = wb["Tray " + str(tray)]
print(ws.title)

c = ws.cell(row = cRow, column = cCol)
c.value = "Picture " + str(pic)

d = ws.cell(row = cRow, column = cellCol)
d.value = result





print(wb.sheetnames)
wb.save(sName)
